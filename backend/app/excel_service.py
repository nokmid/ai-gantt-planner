"""
Импорт и экспорт плана задач в формате Excel.

Ожидаемые колонки на входе (см. пример example_tasks.xlsx):
    задача | описание | исполнитель | длительность | предшественники

"предшественники" — названия других задач через запятую (не id — это
удобнее для человека, заполняющего Excel руками). При импорте мы сами
резолвим имена в id.
"""
from __future__ import annotations

from datetime import date, timedelta
from io import BytesIO

import openpyxl
from openpyxl.utils import get_column_letter

from .models import Task

COLUMNS = ["задача", "описание", "исполнитель", "длительность", "предшественники"]


def parse_excel(file_bytes: bytes) -> list[Task]:
    wb = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wb.active

    header = [str(c.value).strip().lower() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    col_idx = {name: header.index(name) for name in COLUMNS if name in header}
    missing = [c for c in COLUMNS if c not in col_idx]
    if missing:
        raise ValueError(f"В Excel не хватает колонок: {', '.join(missing)}")

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or all(v is None for v in row):
            continue
        rows.append(row)

    # первый проход: создаём задачи с временными id = порядковый номер + имя
    name_to_id: dict[str, str] = {}
    tasks_raw = []
    today = date.today()
    cursor = today
    for i, row in enumerate(rows):
        name = str(row[col_idx["задача"]]).strip()
        description = str(row[col_idx["описание"]] or "").strip()
        assignee = str(row[col_idx["исполнитель"]] or "").strip()
        duration = row[col_idx["длительность"]]
        duration = int(duration) if duration else 1
        preds_raw = row[col_idx["предшественники"]]
        preds_names = [p.strip() for p in str(preds_raw).split(",") if p.strip()] if preds_raw else []

        task_id = f"imp{i}"
        name_to_id[name.lower()] = task_id
        tasks_raw.append({
            "id": task_id,
            "name": name,
            "description": description,
            "assignee": assignee,
            "duration_days": duration,
            "pred_names": preds_names,
        })

    # второй проход: резолвим предшественников по имени -> id, ставим даты
    # даты выставляем эвристически: старт = сегодня, а если есть
    # предшественник — старт = день после окончания последнего предшественника
    tasks: dict[str, Task] = {}

    def resolve(raw) -> Task:
        if raw["id"] in tasks:
            return tasks[raw["id"]]
        pred_ids = [name_to_id[n.lower()] for n in raw["pred_names"] if n.lower() in name_to_id]
        if pred_ids:
            pred_ends = []
            for pid in pred_ids:
                pred_raw = next(r for r in tasks_raw if r["id"] == pid)
                pred_task = resolve(pred_raw)
                pred_ends.append(pred_task.start_date + timedelta(days=pred_task.duration_days))
            start = max(pred_ends)
        else:
            start = today
        task = Task(
            id=raw["id"], name=raw["name"], description=raw["description"],
            assignee=raw["assignee"], start_date=start, duration_days=raw["duration_days"],
            predecessors=pred_ids,
        )
        tasks[raw["id"]] = task
        return task

    for raw in tasks_raw:
        resolve(raw)

    return list(tasks.values())


def export_excel(tasks: list[Task]) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "План"

    ws.append(COLUMNS)
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    id_to_name = {t.id: t.name for t in tasks}
    for t in tasks:
        preds = ", ".join(id_to_name.get(p, p) for p in t.predecessors)
        ws.append([t.name, t.description, t.assignee, t.duration_days, preds])

    for i, col in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(18, len(col) + 4)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
